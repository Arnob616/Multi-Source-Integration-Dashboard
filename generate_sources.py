"""
generate_sources.py  —  Creates realistic Excel + CSV source files
Excel : Monthly sales transactions  (product_code, sales figures, region)
CSV   : Warehouse inventory snapshot (product_code, stock, warehouse, reorder flags)
"""

import random
import pandas as pd
from datetime import datetime, timedelta
from faker import Faker
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

fake = Faker()
random.seed(99)

PRODUCT_CODES = [f"PRD-{1000 + i:04d}" for i in range(80)]
REGIONS       = ["North America", "Europe", "Asia Pacific", "Latin America", "Middle East"]
REPS          = [fake.name() for _ in range(15)]

def _rand_date(days_back=180):
    return (datetime.now() - timedelta(days=random.randint(0, days_back))).strftime("%Y-%m-%d")

def _inject_dirty(df: pd.DataFrame) -> pd.DataFrame:
    """Add realistic data quality issues."""
    df = df.copy()
    # Randomly uppercase some product codes
    mask = df.sample(frac=0.04).index
    df.loc[mask, "product_code"] = df.loc[mask, "product_code"].str.upper().str.replace("-","_")
    # Some negative quantities (returns)
    mask2 = df.sample(frac=0.03).index
    df.loc[mask2, "units_sold"] = -df.loc[mask2, "units_sold"].abs()
    # Some blanks in region
    mask3 = df.sample(frac=0.05).index
    df.loc[mask3, "region"] = None
    # Revenue stored as string in some rows
    mask4 = df.sample(frac=0.04).index
    df["gross_revenue"] = df["gross_revenue"].astype(object)
    df.loc[mask4, "gross_revenue"] = df.loc[mask4, "gross_revenue"].astype(str) + " USD"
    return df

# ── Excel: Sales Data ──────────────────────────────────────────────────────────
def generate_excel(path: str):
    rows = []
    for _ in range(350):
        code  = random.choice(PRODUCT_CODES)
        units = random.randint(1, 500)
        price = round(random.uniform(10, 3000), 2)
        disc  = round(random.uniform(0, 0.30), 2)
        rows.append({
            "transaction_id": f"TXN-{random.randint(100000,999999)}",
            "product_code":   code,
            "sale_date":      _rand_date(180),
            "region":         random.choice(REGIONS),
            "sales_rep":      random.choice(REPS),
            "units_sold":     units,
            "unit_price":     price,
            "discount_pct":   disc,
            "gross_revenue":  round(units * price, 2),
            "net_revenue":    round(units * price * (1 - disc), 2),
            "currency":       "USD",
        })
    df = _inject_dirty(pd.DataFrame(rows))

    # Write base data with pandas
    df.to_excel(path, index=False, sheet_name="Sales_Transactions")

    # Style with openpyxl
    wb = load_workbook(path)
    ws = wb["Sales_Transactions"]

    # Header style
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align= Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="BBBBBB"),
        right =Side(style="thin", color="EEEEEE"),
    )

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align

    # Column widths
    widths = [16, 12, 12, 18, 20, 12, 12, 12, 15, 13, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row alternating fills
    light = PatternFill("solid", start_color="EBF3FB")
    data_font = Font(name="Arial", size=9)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        for cell in row:
            cell.font   = data_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = light

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32
    wb.save(path)
    print(f"  ✓ Excel created: {path}  ({len(df)} rows)")
    return df

# ── CSV: Inventory Data ────────────────────────────────────────────────────────
def generate_csv(path: str):
    rows = []
    warehouses = ["WH-Chicago", "WH-Frankfurt", "WH-Singapore", "WH-Dubai", "WH-SaoPaulo"]
    for code in PRODUCT_CODES:
        for wh in random.sample(warehouses, k=random.randint(1, 3)):
            stock   = random.randint(0, 10000)
            reorder = random.randint(50, 800)
            rows.append({
                "product_code":      code,
                "warehouse":         wh,
                "stock_on_hand":     stock,
                "reserved_stock":    random.randint(0, min(stock, 500)),
                "reorder_point":     reorder,
                "reorder_qty":       reorder * 2,
                "last_counted_date": _rand_date(60),
                "unit_weight_kg":    round(random.uniform(0.1, 50.0), 2),
                "storage_zone":      random.choice(["A", "B", "C", "D"]),
                "below_reorder":     stock < reorder,
                "stock_value_usd":   round(stock * random.uniform(5, 2500), 2),
            })

    df = pd.DataFrame(rows)
    # Inject dirty data
    mask = df.sample(frac=0.05).index
    df.loc[mask, "product_code"] = df.loc[mask, "product_code"].str.lower().str.replace("-", ".")
    mask2 = df.sample(frac=0.04).index
    df.loc[mask2, "stock_on_hand"] = None

    df.to_csv(path, index=False)
    print(f"  ✓ CSV   created: {path}  ({len(df)} rows)")
    return df

if __name__ == "__main__":
    out = Path("data/raw")
    out.mkdir(parents=True, exist_ok=True)
    generate_excel(str(out / "sales_data.xlsx"))
    generate_csv(str(out / "inventory_data.csv"))
    print("Source files ready.")
